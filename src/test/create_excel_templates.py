"""
create_excel_templates.py — generates all required Excel input template files
in the resources/ folder for the AITestCaseGen scripts.

Usage:
    python src/create_excel_templates.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

RESOURCES_DIR = os.path.join(os.path.dirname(__file__), "..", "resources")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="0052CC")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

TEMPLATES = {
    "import_test_case.xlsx": [
        "Folder", "Summary", "Objective", "Status", "Owner",
        "Test Reviewer", "Priority", "Precondition",
        "Test Step", "Test Data", "Expected Result",
        "Traceability", "Type of Test", "Automation Status", "Comments",
    ],
    "update_test_case.xlsx": [
        "Key", "Folder", "Status", "Owner", "Priority",
        "Automation Status", "Test Reviewer", "Test Type",
        "Objective", "Precondition", "Component", "Labels",
        "App Mnemonic", "Traceability",
    ],
    "delete_test_case.xlsx": [
        "Folder", "Test_case", "Name",
    ],
    "update_test_cycle.xlsx": [
        "Key", "Environment", "Status", "Assigned To", "Comment",
    ],
}


def _create_sheet(headers: list, sheet_title: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    return wb


def main():
    os.makedirs(RESOURCES_DIR, exist_ok=True)
    print("=" * 55)
    print("  Creating Excel templates in resources/")
    print("=" * 55)

    for filename, headers in TEMPLATES.items():
        out_path  = os.path.join(RESOURCES_DIR, filename)
        sheet_name = filename.replace(".xlsx", "").replace("_", " ").title()[:31]
        wb = _create_sheet(headers, sheet_name)
        wb.save(out_path)
        print(f"  ✅ {filename}  ({len(headers)} columns)")

    print("=" * 55)
    print(f"  All templates saved to: {os.path.abspath(RESOURCES_DIR)}")
    print("=" * 55)


if __name__ == "__main__":
    main()
