"""
ImportTestCases — reads test cases from resources/import_test_case.xlsx and
creates them in Jira Zephyr Scale.

Expected Excel columns (from import_test_case.xlsx):
    Folder, Summary, Objective, status, owner, Test Reviewer,
    priority, Precondition, Test Step, Test Data, Expected Result,
    Traceability, Type of Test, Automation Status, Comments

Usage:
    python -m test.import_testcases
    python -m test.import_testcases --dry-run
    python -m test.import_testcases --file resources/import_test_case.xlsx
"""

import sys
import os
import ssl
import json
import argparse
import urllib.request
import urllib.error
import urllib.parse

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from main.config import AppConfig
from main.jira_client import JiraClient
from main.report import ReportGenerator


class ImportTestCases:

    DEFAULT_XLSX = os.path.join(
        os.path.dirname(__file__), "..", "..", "resources", "import_test_case.xlsx"
    )

    # Map CSV column name (lowercase, stripped) → Zephyr payload field
    COLUMN_MAP = {
        "summary"          : "name",
        "folder"           : "folder",
        "objective"        : "objective",
        "precondition"     : "precondition",
        "priority"         : "priority",
        "owner"            : "owner",
        "status"           : "status",
        "test reviewer"    : "test_reviewer",
        "type of test"     : "test_type",
        "automation status": "automation_status",
        "traceability"     : "story_link",
        "story link"       : "story_link",
        "issue links"      : "story_link",
        "test step"        : "test_step",
        "test data"        : "test_data",
        "expected result"  : "expected_result",
        "comments"         : "comments",
    }

    def __init__(self, xlsx_path: str = "", dry_run: bool = False):
        self._cfg       = AppConfig()
        self._client    = JiraClient()
        self.xlsx_path  = xlsx_path or self.DEFAULT_XLSX
        self.dry_run    = dry_run

    # — Read ————————————————————————————————————————————————————————

    def load_xlsx(self) -> list:
        """Read Excel and return list of normalised row dicts."""
        if not os.path.exists(self.xlsx_path):
            print(f"[ERROR] Excel file not found: {os.path.abspath(self.xlsx_path)}")
            sys.exit(1)

        wb = openpyxl.load_workbook(self.xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        raw_headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"  Detected columns : {raw_headers}")
        header_lookup = {h.lower().split("\n")[0].strip(): idx for idx, h in enumerate(raw_headers)}

        if "summary" not in header_lookup:
            print(f"[ERROR] 'Summary' column not found. Headers: {raw_headers}")
            sys.exit(1)

        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name_idx = header_lookup.get("summary")
            name = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else ""
            if not name or name == "None":
                continue
            record = {"_row": i}
            for col_lower, field in self.COLUMN_MAP.items():
                idx = header_lookup.get(col_lower)
                value = str(row[idx]).strip() if idx is not None and row[idx] is not None else ""
                if value == "None":
                    value = ""
                if value or field not in record:
                    record[field] = value
            rows.append(record)

        wb.close()
        print(f"  Loaded {len(rows)} test case(s) from: {os.path.basename(self.xlsx_path)}\n")
        return rows

    # — Build payload ————————————————————————————————————————————————

    def build_payload(self, record: dict) -> dict:
        """Construct the Zephyr Scale POST payload from a CSV row."""
        cfg = self._cfg

        payload = {
            "projectKey" : cfg.project_key,
            "name"       : record["name"],
            "status"     : record.get("status") or "Approved",       # default: Approved
            "priority"   : record.get("priority") or "Medium",       # default: Medium
            "folder"     : (record.get("folder").strip().rstrip("/") if record.get("folder") else "/"),  # default: root /
        }

        # owner: use config owner_account_id always (most reliable)
        payload["owner"] = cfg.owner_account_id

        # Only include customFields that have a non-empty value
        custom_fields = {}
        # Test Reviewer: config account ID takes precedence over CSV value
        test_reviewer = cfg.test_reviewer_id if cfg.test_reviewer_id else (record.get("test_reviewer") or "")
        if test_reviewer:
            custom_fields["Test Reviewer"] = test_reviewer
        if record.get("test_type"):
            custom_fields["Test Type"] = record["test_type"]
        if record.get("automation_status"):
            custom_fields["Automation Status"] = record["automation_status"]
        if custom_fields:
            payload["customFields"] = custom_fields

        if record.get("objective"):
            payload["objective"] = record["objective"]

        if record.get("precondition"):
            payload["precondition"] = record["precondition"]

        story_link = record.get("story_link", "").strip()
        if story_link:
            # Support comma or space separated keys e.g. "KEY-1, KEY-2" or "KEY-1 KEY-2"
            separators = story_link.replace(",", " ")
            payload["issueLinks"] = [l.strip() for l in separators.split() if l.strip()]

        # Build test script — all steps as a single step block
        step_block = record.get("test_step", "")
        test_data  = record.get("test_data", "")
        expected   = record.get("expected_result", "")

        if step_block:
            payload["testScript"] = {
                "type" : "STEP_BY_STEP",
                "steps": [
                    {
                        "description"   : step_block.strip(),
                        "testData"      : test_data,
                        "expectedResult": expected,
                    }
                ],
            }

        return payload

    # — Resolve owner ————————————————————————————————————————————————

    def resolve_owner(self, value: str) -> str:
        """Return account ID for the given value.
        If value has no spaces it is already an account ID — return as-is.
        If it looks like a display name, query Jira user search API.
        Falls back to config owner_account_id on failure.
        """
        if " " not in value:
            return value  # already an account ID

        cfg = self._cfg
        url = cfg.base_url + "/rest/api/2/user/search?username=" + urllib.parse.quote(value.split()[0])
        req = urllib.request.Request(url, headers={
            "Authorization": "Bearer " + cfg.jira_token,
            "Accept"       : "application/json",
        })
        ctx = ssl.create_default_context()
        try:
            with urllib.request.urlopen(req, context=ctx) as r:
                users = json.loads(r.read())
                for u in users:
                    if u.get("displayName", "").lower() == value.lower():
                        account_id = u.get("name") or u.get("accountId", "")
                        print(f"\n    [INFO] Resolved owner '{value}' + {account_id}", end="")
                        return account_id
        except Exception:
            pass
        print(f"\n    [WARN] Could not resolve owner '{value}' — using config default.", end="")
        return cfg.owner_account_id

    # — Create ————————————————————————————————————————————————————————

    def _post_jira_comment(self, issue_key: str, comment: str):
        """POST comment to Jira issue via REST API."""
        cfg = self._cfg
        url = f"{cfg.base_url}/rest/api/2/issue/{issue_key}/comment"
        body = json.dumps({"body": comment}).encode("utf-8")
        req = urllib.request.Request(url, data=body, headers={
            "Authorization": f"Bearer {cfg.jira_token}",
            "Content-Type" : "application/json",
        }, method="POST")
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode    = ssl.CERT_NONE
        try:
            with urllib.request.urlopen(req, context=ctx) as r:
                pass
        except urllib.error.HTTPError:
            pass

    def create_test_case(self, payload: dict, comment: str = "") -> tuple:
        """POST /testcase then optionally add a comment. Returns (success, key_or_error)."""
        status, data = self._client.post("/testcase", payload)
        if status in (200, 201):
            key = data.get("key", "?")
            if comment:
                self._post_jira_comment(key, comment)
            return True, key
        raw = data.get("error", "")
        if raw and not data.get("errorMessages"):
            try:
                data = json.loads(raw)
            except Exception:
                pass
        errors = data.get("errorMessages", [])
        messages = []
        for msg in errors:
            if "folder" in msg.lower():
                messages.append(
                    f"Folder '{payload.get('folder')}' does not exist in project '{payload.get('projectKey')}'. "
                    f"Please check the folder path and project name in your CSV."
                )
            elif "owner" in msg.lower() or "user" in msg.lower():
                messages.append(
                    f"Owner '{payload.get('owner')}' was not found in Jira. Check the account ID or update owner_account_id in config.properties."
                )
            else:
                messages.append(msg)
        return False, "; ".join(messages) if messages else f"{status} — {data}"

    # — Run ————————————————————————————————————————————————————————————

    def run(self):
        print("=" * 65)
        print("  Jira Zephyr Scale — Import Test Cases from Excel")
        print("=" * 65)
        print(f"  Project  : {self._cfg.project_key}")
        print(f"  Excel File: {os.path.abspath(self.xlsx_path)}")
        if self.dry_run:
            print("  Mode     : DRY RUN (no changes will be made)")
        print("=" * 65 + "\n")

        records = self.load_xlsx()

        if not records:
            print("[WARN] No valid rows found in Excel file.")
            sys.exit(0)

        ok_list, fail_list = [], []
        test_cases = []  # [{key, name}] for report

        for i, record in enumerate(records, 1):
            name    = record["name"]
            payload = self.build_payload(record)

            print(f"  [{i:>3}/{len(records)}] {name[:60]:<60s} ... ", end="", flush=True)

            if self.dry_run:
                print("SKIP (dry-run)")
                test_cases.append({"key": f"ROW-{i}", "name": name})
                continue

            success, result = self.create_test_case(payload, comment=record.get("comments", ""))
            if success:
                print(f"✅  {result}")
                ok_list.append(result)
                test_cases.append({"key": result, "name": name})
            else:
                print(f"❌  {result}")
                fail_list.append(name)
                test_cases.append({"key": name, "name": name})

        print()
        print("=" * 65)
        if self.dry_run:
            print(f"  DRY RUN — {len(records)} test case(s) would be created.")
        else:
            print(f"  SUMMARY")
            print(f"  Created : {len(ok_list)}")
            print(f"  Failed  : {len(fail_list)}")
            if ok_list:
                print(f"\n  Created keys: {', '.join(ok_list)}")
            if fail_list:
                print(f"\n  Failed:")
                for name in fail_list:
                    print(f"    — {name}")
        print("=" * 65)

        if not self.dry_run:
            cfg = self._cfg
            reports_dir = os.path.join(os.path.dirname(__file__), "..", "..", "reports")
            os.makedirs(reports_dir, exist_ok=True)
            report_path = os.path.join(reports_dir, "Import_TC_Report.html")
            folder_path = records[0].get("folder", "/") if records else "/"
            ReportGenerator(output_path=report_path, action_label="Created", report_title="Zephyr Scale — Add Test Cases Report").generate(
                test_cases=test_cases,
                ok_list=ok_list,
                fail_list=fail_list,
                project_key=cfg.project_key,
                folder_path=folder_path,
            )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import test cases from CSV into Jira Zephyr Scale")
    parser.add_argument("--file",    default="", help="Path to Excel file (default: resources/import_test_case.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Preview without creating")
    args = parser.parse_args()
    ImportTestCases(xlsx_path=args.file, dry_run=args.dry_run).run()
